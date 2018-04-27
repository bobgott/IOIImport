import argparse
import configparser
import datetime
import logging
import os
import sys
import openpyxl
from applic.dicttoxml import DictToXML, DXMLGeneratedError
from files_folders import FilesAndFolders

__project__ = 'IOI_Import'
__author__ = "Robert Gottesman"
__version_date__ = "04/24/2018"
__err_prefix__ = 'IOI'
__high_err_num__ = 13

""" Change Log
04/17/2017 - Groups column can be separated with '_' or ','
06/14/2017 - Removed DD Wiki Version text (not used)
10/29/2017 - Program starting from __init()__.py. Changed main folder name to 'files'
04/21/2018 - Added Spanish columns to config.ini
"""


class IOIGeneratedError(Exception):
    """
    Handle known problems in this module passing detail information
    """
    def __init__(self, value):
        self.value = value

    def __str__(self):
        return repr(self.value)


def valid_date(date_string):
    """ Return date/time object in YYYY-MM-DD format

    :param date_string: (str) Date in YYYY-MM-DD
    :return: date object
    """
    try:
        return datetime.datetime.strptime(date_string, "%Y-%m-%d")
    except ValueError:
        msg = "Not a valid date: '{0}'.".format(date_string)
        raise argparse.ArgumentTypeError(msg)


def print_lookup_fields(spreadsheet_info):
    """ Used for debugging (ignore)

    :param spreadsheet_info: Dictionary format of xlsx/csv file
    :return: None
    """
    top_index = spreadsheet_info['Lookups']
    for letter in sorted(top_index.keys()):
        print('{0} - Lookup Fields'.format(letter))
        # note: lookup_fields['PropertySubType'][0]['Enumeration'] .. get Enumeration
        for fld_name in top_index[letter]:
            print(fld_name)


class ResoXLSXtoDict:

    def __init__(self, config_file_path, xlsx_filepath):
        """ Read xlsx files into internal dictionary 'spreadsheet_info'

        :param config_file_path: (str) Full path for config.ini
        :param xlsx_filepath: (str) Full path for input xlsx file
        :return: Void. Raise IOIGeneratedError on error
        """
        self.xlsx_filepath = xlsx_filepath
        self.logger = logging.getLogger(__project__ + '.' + self.__class__.__name__)
        self.logger.debug("Initialize {0} with verson:{1}".format(self.__class__.__name__, __version_date__))
        self.resource_sheets = []
        self.lookup_sheet = None
        self.spreadsheet_info = {'Resources': {}, 'Lookups': {}}  # Container of xlsx data
        self._read_config_ini(config_file_path)

    def _read_config_ini(self, config_file_path):
        """ Read in program configuration (config.ini). See readme.md for detail

        :param config_file_path: (str) config.ini filepath
        :return: Void. Raise IOIGeneratedError on error.
        """
        # See: https://docs.python.org/3/library/configparser.html
        self.config = configparser.ConfigParser()
        self.config.optionxform = str   # Setting str, makes option names case sensitive:
        if not os.path.isfile(config_file_path):
            raise IOIGeneratedError("[IOI-06] Cannot Process Program Config INI File: " + config_file_path)
        try:
            self.config.read(config_file_path)
        except FileNotFoundError:
            raise IOIGeneratedError("[IOI-01] Cannot Process Program Config INI File: " + config_file_path)

        # Read in Resource Sheets, formatted as: [sheet tab name] = [output xml resource name]
        try:
            if 'ResourceSheets' in self.config and len(self.config['ResourceSheets']) > 0:
                resource_sheets_found = True
                for resource_sheet in self.config['ResourceSheets']:
                    self.resource_sheets.append(resource_sheet)
            else:
                resource_sheets_found = False
        except KeyError:
            raise IOIGeneratedError("[IOI-12] Cannot find 'ResourceSheets' key in config.ini")

        # Get Lookup Sheet Name (only 1 allowed), formatted as: [LookupSheet] = [lookup spreadsheet tab name]
        try:
            if 'LookupSheets' in self.config and len(self.config['LookupSheets']) > 0:
                lookup_sheets_found = True
                self.lookup_sheet = self.config['LookupSheets']['LookupSheet']
            else:
                lookup_sheets_found = False
        except KeyError:
            raise IOIGeneratedError("[IOI-09] Cannot find 'LookupSheet' or child 'LookupSheets' key config.ini")

        if not len(self.resource_sheets) == 0 and self.lookup_sheet is None:
            raise IOIGeneratedError("[IOI-13] Missing entries in [ResourceSheets] and [LookupSheets] in config.ini")

    def read_xlsx_file(self):
        """ Open .xlsx file and read into self.spreadsheet_info

        :return: void. Raise IOIGeneratedError on error
        """
        try:
            wb = openpyxl.load_workbook(self.xlsx_filepath)
        except FileNotFoundError:
            raise IOIGeneratedError('[IOI-07] XLSX input file {0} not found'.format(self.xlsx_filepath))
        # Read in Resource Sheet rows
        for resource_sheet_name in self.resource_sheets:
            try:
                ws = wb.get_sheet_by_name(name=resource_sheet_name)
            except KeyError:
                raise IOIGeneratedError("[IOI-11] Resource Sheet name '{0}' does not exist in .xlsx file".
                                        format(resource_sheet_name))
            self.logger.info("Reading Input Resource Worksheet: '{}'".format(ws.title))
            self._create_resource_dict(resource_sheet_name, ws) # fill self.spreadsheet_info['Resources'][sheet_name]
        # Read in Lookup Sheet rows
        if self.lookup_sheet is not None:
            try:
                ws = wb.get_sheet_by_name(name=self.lookup_sheet)
            except KeyError:
                raise IOIGeneratedError("[IOI-08] Lookup Sheet name '{0}' does not exist in .xlsx file".
                                        format(self.lookup_sheet))
            self.logger.info("Reading Input Lookup Worksheet: '{}'".format(ws.title))
            self._create_lookup_dict(ws)  # Fill in self.spreadsheet_info['Lookups']

            if len(self.spreadsheet_info['Lookups']) == 0:
                raise IOIGeneratedError('[W202] No Lookup Lookups Processed (tab: {})'.format(self.lookup_sheet))


    def _create_lookup_dict(self, ws):
        """ Populate xlsx lookup rows into internal dictionary (self.spreadsheet_info['Lookups'])

        :param ws: (obj) lookup xlsx worksheet object
        :return: void. Raise IOIGeneratedError on error
        """
        # Each key in lookup_fiels a field name, the value is a list of lookup values
        # .. (Example 'PropertySubType Lookups - see: http://ddwiki.reso.org/display/DDW/PropertySubType+Lookups)
        lookup_fields = {}

        header_cols = [ws.cell(row=1, column=idx).value for idx in range(1, ws.max_column+1)
                       if ws.cell(row=1, column=idx).value is not None]

        for row in range(2, ws.max_row + 1):
            # Each entry in lookup_field is a lookup field. Value is a list of lookup values
            lookup_fields = self.fillin_lookupfield_byrow(ws, lookup_fields, header_cols, row)

        # Loop through every lookup field and create entry in top_index {}
        for fld in lookup_fields:
            if fld is not None:
                # Each key in dict self.spreadsheet_info['Lookups'] is a single letter'
                # .. The value is a list of lookup field names. This mimics DD Wiki
                # .. (Example 'A' - see: http://ddwiki.reso.org/display/DDW/A+-+Lookup+Fields)
                self.spreadsheet_info['Lookups'].setdefault(fld[0], []).append([fld, lookup_fields[fld]])

    def fillin_lookupfield_byrow(self, ws, lookup_fields, header_cols, row):
        """ Read row from spreadsheet and translate to internal dictionary object

        :param ws: (obj) Worksheet object
        :param lookup_fields: (dict) Partial Container for all lookup fields and values
        :param header_cols: (list) All header columns
        :param row: (int) Row being submitted
        :return: (dict) Container for all lookup fields and values. Raise IOIGeneratedError on error.
        """
        my_row = {}
        for col_num, col_val in enumerate(header_cols):
            my_row[col_val] = ws.cell(row=row, column=col_num+1).value
        try:
            lookup_fields.setdefault(my_row['LookupField'], []).append(my_row)
        except KeyError:
            raise IOIGeneratedError("[IOI-10] Cannot find field 'LookupField' in spreadsheet")
        return lookup_fields

    def _create_resource_dict(self, sheet_tab_name, ws):
        """ Populate xlsx resource/collection rows into internal dictionary (self.spreadsheet_info['Resources'])

        :param ws: (obj) lookup xlsx worksheet object
        :return: void. Raise IOIGeneratedError on error
        """
        self.spreadsheet_info['Resources'][sheet_tab_name] = {}

        header_cols = [ws.cell(row=1, column=idx).value for idx in range(1, ws.max_column+1)
                       if ws.cell(row=1, column=idx).value is not None]

        for row in range(2, ws.max_row + 1):
            my_row = {}
            if ws.cell(row=row, column=1).value is not None and len(ws.cell(row=row, column=1).value) > 0:
                for col_num, col_val in enumerate(header_cols):
                    my_row[col_val] = ws.cell(row=row, column=col_num + 1).value
                self._replace_val_in_groups(my_row)  # Replace string with list
                self.spreadsheet_info['Resources'][sheet_tab_name].setdefault(','.join(my_row["Groups"]), []).\
                    append(my_row)

    def _replace_val_in_groups(self, my_row):
        """ Create a list value for entry 'Groups' column in xlsx row (not to be confused with 'Groups' output xml tag)
        
        :param my_row: (obj) A row from the input .xlsx file
        :return: Null
        """
        my_row['Groups'] = [y for y in [x.strip() for x in my_row['Groups'].split(',')] if y]


def main(argv):
    # https://docs.python.org/3.3/library/argparse.html
    # https://docs.python.org/3/howto/argparse.html
    parser = argparse.ArgumentParser(description='RESO xlsx to xml Import')
    parser.add_argument('-f', '--home_folder', default=None,
                        help="Default folder for config, ini and error log files <current folder>")
    parser.add_argument('-c', '--config_sub_folder', default='current',
                        help="Sub Folder in files/config containg ini files <current>")
    parser.add_argument('-x', '--xlsx_filename', default=None,
                        help="Input .xlsx file containing new fields and lookups")
    parser.add_argument('-i', '--max_id_filename', default='stat_warning_log.txt',
                        help="Input file containing DD Wiki max record and lookup ids <stat_warning_log.txt>")
    parser.add_argument('-w', '--ddwiki_exported_xml_filename', default=None,
                        help="Input Previous Confluence DD Wiki exported XML file to check for duplicate page titles")
    # type=valid_date .. validate date input with valid_date()
    parser.add_argument('-d', '--xlsx_date', type=valid_date, default=None,
                        help="Noted create date for input spreadsheet (YYYY-MM-DD)")
    parser.add_argument('-e', '--error_logging', type=int, default=20,
                        help="Error Logging Level (0-None, 10-Debug, 20-Info, 30-Warn, 40-Err, 50-Critical <20>")
    args = parser.parse_args()

    faf = FilesAndFolders(args.home_folder, args.config_sub_folder) if args.home_folder else \
        FilesAndFolders(os.getcwd(), args.config_sub_folder)

    # xlsx create date (in xml root node) - Enter manually based on DD Spreadsheet
    if args.xlsx_date is None:
        xlsx_date = datetime.datetime.now()
    else:
        xlsx_date = args.xlsx_date
    logger = logging.getLogger(__project__)
    logger.setLevel(args.error_logging)
    formatter = logging.Formatter('%(asctime)s (%(name)s:%(levelname)s) %(message)s')
    stream_display = logging.StreamHandler()
    stream_display.setFormatter(formatter)
    stream_file = logging.FileHandler(os.path.join(faf.log_folder, "logging " +
                                                   datetime.datetime.today().strftime('%Y-%m-%d') + '.txt'), mode='w')
    stream_file.setFormatter(formatter)
    logger.addHandler(stream_display)
    logger.addHandler(stream_file)
    logger.info("Starting IOI xlsx-to-xml. Program verson:{0}".format(__version_date__))

    input_xlsx_filepath = os.path.join(faf.input_folder, args.xlsx_filename)
    # max_id_filename - file created by RESOExporter. Contains max rec/lookup ids (i.e ddwiki_stat_log2017-05-12.txt)
    max_id_filepath = os.path.join(faf.input_folder, args.max_id_filename)
    ddwiki_exported_filepath = os.path.join(faf.input_folder, args.ddwiki_exported_xml_filename)

    # Read in csv files and convert to internal python dict {}
    logger.info("Importing file:'{}' with date:{}".format(args.xlsx_filename, xlsx_date.strftime('%m-%d-%Y %H:%M')))
    logger.info("Base Folder for config files is: " + args.config_sub_folder)
    logger.info("Base Data File Input Folder is: input")
    logger.info("Input RESO Export XML file:{}".format(args.ddwiki_exported_xml_filename))
    logger.info("Input RESO Stat/Max ID File:{}".format(args.max_id_filename))
    try:
        # Create object to convert xlsx into xml
        xlsx_to_dict = ResoXLSXtoDict(config_file_path=faf.config_file,
                                      xlsx_filepath=input_xlsx_filepath)
    except IOIGeneratedError as e:
        logger.error("? Error initiating ResoXLSXtoDict: " + e.value)
        sys.exit(-1)
    # result_xml_filename = os.path.basename(input_xlsx_filepath)
    faf.xml_filepath = os.path.splitext(os.path.basename(input_xlsx_filepath))[0] + '.xml'
    logger.info("Resultant Output IOI XML File: " + faf.xml_filepath)

    try:
        # Read xlsx into internal structure xlsx_to_dict.spreadsheet_info
        xlsx_to_dict.read_xlsx_file()
    except IOIGeneratedError as e:
        logger.error("Error reading .xlsx file: " + e.value)
        sys.exit(-1)
    try:
        # Convert internal structure into IOI xml file
        DictToXML(files_and_folders=faf, result_xml_filepath=faf.xml_filepath, max_id_filepath=max_id_filepath,
                  ddwiki_exported_filepath=ddwiki_exported_filepath, spreadsheet_dict=xlsx_to_dict.spreadsheet_info,
                  xlsx_date=xlsx_date, program_config_data=xlsx_to_dict.config)
    except DXMLGeneratedError as e:
        logger.error("Error creating XML File: " + e.value)
        sys.exit(-1)

    logger.info('** Program Ends in Success **')

if __name__ == "__main__":
    main(sys.argv[1:])

# Low: ?? See References column in lookup page 'Agent (NotedBy)'. Collection concatenated (https://goo.gl/0pj92s)

